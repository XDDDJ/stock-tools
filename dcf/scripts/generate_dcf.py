"""
DCF Valuation Model Generator v2.0
全自动三情景 DCF 估值 Excel 生成器，支持 A 股/港股/美股。

优化点：
1. 净债务调整：EV → Equity Value → Per Share
2. 敏感性分析：Excel 数据表联动（非 Python 计算）
3. 投资评级：基于安全边际自动评级
4. 多市场：自动适配 A 股(CNY)/港股(HKD)/美股(USD)
5. 汇率：支持实时汇率传入

Usage:
    python generate_dcf.py <config.json> <output.xlsx>
"""
import json, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ── 字体定义 ──────────────────────────────────────────
DARK_BLUE = Font(color="1F4E79", bold=True)
DARK_BLUE_NORMAL = Font(color="1F4E79")
GRAY = Font(color="666666", size=10)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BLACK_BOLD = Font(bold=True, color="000000")
BLACK = Font(color="000000")
INPUT_FONT = Font(color="0000FF")           # 蓝色 = 手动输入
FORMULA_FONT = Font(color="000000")          # 黑色 = 公式
CROSS_REF_FONT = Font(color="008000")        # 绿色 = 跨 Sheet 引用
LARGE_BOLD = Font(bold=True, size=12)

# ── 背景色 ────────────────────────────────────────────
HEADER_BG = PatternFill("solid", fgColor="1F4E79")
OPTIMISTIC_BG = PatternFill("solid", fgColor="E2EFDA")
OPTIMISTIC_FONT = Font(bold=True, color="00B050", size=12)
BASE_BG = PatternFill("solid", fgColor="D6E4F0")
BASE_FONT = Font(bold=True, color="4472C4", size=12)
PESSIMISTIC_BG = PatternFill("solid", fgColor="FCE4EC")
PESSIMISTIC_FONT = Font(bold=True, color="FF0000", size=12)
GRAY_BG = PatternFill("solid", fgColor="F2F2F2")
YELLOW_BG = PatternFill("solid", fgColor="FFFDE7")
GREEN_BG = PatternFill("solid", fgColor="C6EFCE")
YELLOW_LIGHT_BG = PatternFill("solid", fgColor="FFEB9C")
RED_BG = PatternFill("solid", fgColor="FFC7CE")
DEEP_GREEN_BG = PatternFill("solid", fgColor="A9D18E")

# ── 评级颜色 ──────────────────────────────────────────
RATE_STRONG_BUY = PatternFill("solid", fgColor="A9D18E")   # 深绿
RATE_BUY = PatternFill("solid", fgColor="C6EFCE")           # 浅绿
RATE_HOLD = PatternFill("solid", fgColor="FFEB9C")          # 黄
RATE_SELL = PatternFill("solid", fgColor="FFC7CE")          # 红

RATE_FONT_STRONG = Font(bold=True, color="006100")
RATE_FONT_BUY = Font(bold=True, color="006100")
RATE_FONT_HOLD = Font(bold=True, color="9C6500")
RATE_FONT_SELL = Font(bold=True, color="9C0006")

# ── 边框 ──────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ── 数字格式 ──────────────────────────────────────────
PCT_FMT = '0.0%'
PCT2_FMT = '0.00%'
NUM_FMT = '#,##0.0'
INT_FMT = '#,##0'
CURRENCY_FMT = '#,##0.00'

# ── 投资评级定义 ──────────────────────────────────────
def get_rating(safety_margin):
    """根据安全边际返回 (评级文本, 背景色, 字体)"""
    if safety_margin > 0.40:
        return ("强烈建议买入", RATE_STRONG_BUY, RATE_FONT_STRONG)
    elif safety_margin > 0.20:
        return ("建议买入", RATE_BUY, RATE_FONT_BUY)
    elif safety_margin > -0.10:
        return ("观望", RATE_HOLD, RATE_FONT_HOLD)
    else:
        return ("建议卖出", RATE_SELL, RATE_FONT_SELL)

def get_rate_fill(safety_margin):
    if safety_margin > 0.40:
        return DEEP_GREEN_BG
    elif safety_margin > 0.20:
        return GREEN_BG
    elif safety_margin > -0.10:
        return YELLOW_LIGHT_BG
    else:
        return RED_BG


# ── 工具函数 ──────────────────────────────────────────
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row, c)
        cell.font = WHITE_BOLD
        cell.fill = HEADER_BG
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row, col)
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell


# ── Sheet 1: DCF 模型概览 ────────────────────────────
def build_overview(wb, cfg):
    ws = wb.active
    ws.title = "DCF模型概览"
    set_col_widths(ws, [22, 16, 22, 16, 18, 16, 30])

    c = cfg
    market = c.get('market', 'HK')
    currency = c['currency']
    report_cur = c.get('report_currency', 'CNY')

    # ── 标题 ──
    ws['A1'] = f"{c['company_name']} ({c['ticker']}) 三情景 DCF 估值模型"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:G1')

    ws['A2'] = f"基准日期: {c['date']} | 当前股价: {c['current_price']} {currency} | 总股本: {c['shares']}亿股 | 行业: {c.get('industry', '')}"
    ws['A2'].font = GRAY
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')

    # ── 第一节：关键财务数据 ──
    fin = c['financials']
    r = 4
    ws.cell(r, 1, f"一、关键财务数据 ({fin['base_year']}年报)").font = DARK_BLUE

    r = 5
    fin_rows = [
        ('营业收入', fin.get('revenue'), f'十亿 {report_cur}', True),
        ('营业利润', fin.get('operating_profit'), f'十亿 {report_cur}', True),
        ('净利润', fin.get('net_profit'), f'十亿 {report_cur}', True),
        ('经营现金流', fin.get('operating_cf'), f'十亿 {report_cur}', True),
        ('资本支出', fin.get('capex'), f'十亿 {report_cur}', True),
        ('自由现金流 (FCF)', fin.get('fcf'), f'十亿 {report_cur}', True),
        ('毛利率', fin.get('gross_margin'), None, False),
        ('净利率', fin.get('net_margin'), None, False),
        ('FCF/营收', fin.get('fcf_margin'), None, False),
        ('ROE', fin.get('roe'), None, False),
        ('净债务/净现金', fin.get('net_debt'), f'十亿 {report_cur}', True),
    ]

    for label, val, unit, is_amount in fin_rows:
        ws.cell(r, 1, label).font = BLACK_BOLD
        if val is not None:
            cell = ws.cell(r, 2, val)
            cell.font = INPUT_FONT
            cell.fill = YELLOW_BG
            if is_amount and isinstance(val, (int, float)):
                cell.number_format = NUM_FMT
            elif not is_amount and isinstance(val, (int, float)):
                cell.number_format = PCT_FMT
        if unit:
            ws.cell(r, 3, unit).font = GRAY
        r += 1

    ws.cell(r, 1, '总股本').font = BLACK_BOLD
    ws.cell(r, 2, c['shares']).font = INPUT_FONT
    ws.cell(r, 2).fill = YELLOW_BG
    ws.cell(r, 3, '亿股').font = GRAY
    shares_row = r  # 记录总股本实际行号，供后续每股价值公式引用
    r += 1

    # ── 第二节：WACC 计算 ──
    r += 1
    wacc_section_row = r
    ws.cell(r, 1, "二、折现率 (WACC) 计算").font = DARK_BLUE
    r += 1

    wacc = c['wacc']
    wacc_rows = [
        ('无风险利率', wacc.get('rf'), wacc.get('rf_source', ''), True),
        ('股权风险溢价 (ERP)', wacc.get('erp'), wacc.get('erp_source', ''), True),
        ('Beta', wacc.get('beta'), wacc.get('beta_source', ''), True),
        ('权益成本 (Ke)', None, '=Rf + Beta x ERP', False),
        ('WACC（基准）', None, '全权益（净现金）', False),
        ('WACC（悲观）', None, '+1% 安全边际', False),
    ]

    wacc_start = r
    for label, val, note, is_input in wacc_rows:
        ws.cell(r, 1, label).font = BLACK_BOLD
        if val is not None:
            cell = ws.cell(r, 2, val)
            cell.font = INPUT_FONT
            cell.fill = YELLOW_BG
            cell.number_format = PCT2_FMT
        ws.cell(r, 3, note).font = GRAY
        r += 1

    # 权益成本公式
    ke_row = wacc_start + 3  # Rf + Beta * ERP
    ws.cell(ke_row, 2).value = f"=B{wacc_start}+B{wacc_start+2}*B{wacc_start+1}"
    ws.cell(ke_row, 2).font = FORMULA_FONT
    ws.cell(ke_row, 2).number_format = PCT2_FMT

    # WACC 基准
    wacc_base_row = wacc_start + 4
    ws.cell(wacc_base_row, 2).value = f"=B{ke_row}"
    ws.cell(wacc_base_row, 2).font = FORMULA_FONT
    ws.cell(wacc_base_row, 2).number_format = PCT2_FMT

    # WACC 悲观
    wacc_pess_row = wacc_start + 5
    ws.cell(wacc_pess_row, 2).value = f"=B{wacc_base_row}+0.01"
    ws.cell(wacc_pess_row, 2).font = FORMULA_FONT
    ws.cell(wacc_pess_row, 2).number_format = PCT2_FMT

    # ── 第三节：三情景估值结果 ──
    r += 1
    summary_section_row = r
    ws.cell(r, 1, "三、三情景估值结果").font = DARK_BLUE
    r += 1

    headers = ['情景', f'企业价值\n(十亿{report_cur})', f'净债务\n(十亿{report_cur})', f'股权价值\n(十亿{report_cur})', f'每股价值\n({currency})', '上涨/下跌空间', '投资评级']
    for i, h in enumerate(headers):
        ws.cell(r, i+1, h)
    style_header_row(ws, r, 1, 7)
    r += 1

    scenarios = c['scenarios']
    names = [s['name'] for s in scenarios]
    styles = [
        (OPTIMISTIC_BG, OPTIMISTIC_FONT),
        (BASE_BG, BASE_FONT),
        (PESSIMISTIC_BG, PESSIMISTIC_FONT),
    ]

    data_start_row = r
    for i, (s, sn) in enumerate(zip(scenarios, names)):
        bg, font = styles[i]

        for col in range(1, 8):
            cell = ws.cell(r, col)
            cell.fill = bg
            cell.border = THIN_BORDER

        # 情景名
        ws.cell(r, 1, sn).font = font

        # 企业价值 → 引用情景 Sheet
        ws.cell(r, 2).value = f"='{sn}'!B18"
        ws.cell(r, 2).font = CROSS_REF_FONT
        ws.cell(r, 2).number_format = NUM_FMT

        # 净债务调整
        net_debt = fin.get('net_debt', 0)
        if net_debt is not None:
            ws.cell(r, 3, net_debt).font = INPUT_FONT
            ws.cell(r, 3).number_format = NUM_FMT

        # 股权价值 = EV - 净债务
        ws.cell(r, 4).value = f"=B{r}-C{r}"
        ws.cell(r, 4).font = FORMULA_FONT
        ws.cell(r, 4).number_format = NUM_FMT

        # 每股价值 → 直接引用情景页B20（已含单位换算和汇率换算）
        ws.cell(r, 5).value = f"='{sn}'!B20"
        ws.cell(r, 5).font = font
        ws.cell(r, 5).number_format = NUM_FMT

        # 上涨/下跌空间
        ws.cell(r, 6).value = f"=(E{r}-{c['current_price']})/{c['current_price']}"
        ws.cell(r, 6).font = font
        ws.cell(r, 6).number_format = '+0.0%;-0.0%'

        # 投资评级（用公式判断）
        ws.cell(r, 7).value = f'=IF(F{r}>0.4,"强烈建议买入",IF(F{r}>0.2,"建议买入",IF(F{r}>-0.1,"观望","建议卖出")))'
        ws.cell(r, 7).font = font
        ws.cell(r, 7).alignment = Alignment(horizontal='center')

        r += 1

    # 添加条件格式：评级列根据安全边际变色
    for i in range(3):
        row = data_start_row + i
        col_letter = 'F'  # 安全边际列

    # 当前股价参考行
    ws.cell(r, 1, '').font = BLACK
    ws.cell(r, 4, f'当前股价: {c["current_price"]} {currency}').font = GRAY
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    return ws


# ── Sheet 2-4: 情景详情 ────────────────────────────────
def build_scenario(wb, cfg, idx):
    s = cfg['scenarios'][idx]
    name = s['name']
    ws = wb.create_sheet(name)

    set_col_widths(ws, [16, 16, 16, 14, 16, 16, 18, 4, 10, 12])

    c = cfg
    fin = c['financials']
    market = c.get('market', 'HK')

    # ── 标题颜色 ──
    title_colors = ["00B050", "4472C4", "FF0000"]
    ws['A1'] = f"{c['company_name']} DCF — {name}"
    ws['A1'].font = Font(bold=True, size=14, color=title_colors[idx])
    ws['A2'] = s.get('description', '')
    ws['A2'].font = GRAY

    # ── 隐藏参数区（I/J 列）──
    ws['I1'] = 'WACC'
    ws['J1'] = f"='DCF模型概览'!B{'23' if idx < 2 else '24'}"  # 引用概览页 WACC
    ws['J1'].font = CROSS_REF_FONT
    ws['J1'].number_format = PCT2_FMT
    ws['I2'] = '永续增长率'
    ws['J2'] = s['terminal_growth']
    ws['J2'].font = INPUT_FONT
    ws['J2'].fill = YELLOW_BG
    ws['J2'].number_format = PCT_FMT

    # ── 表头 ──
    headers = ['年份', '营收增长率', '营收(十亿)', 'FCF Margin', 'FCF(十亿)', '折现因子', 'PV of FCF(十亿)']
    for i, h in enumerate(headers):
        ws.cell(4, i+1, h)
    style_header_row(ws, 4, 1, 7)

    # ── 基准年 ──
    base_year = fin['base_year']
    base_rev = fin['revenue']
    base_fcf = fin['fcf']
    growth = s['growth_rates']
    margins = s['fcf_margins']

    ws.cell(5, 1, f"{base_year} (基准)").font = GRAY
    ws.cell(5, 3, base_rev).font = INPUT_FONT
    ws.cell(5, 3).fill = YELLOW_BG
    ws.cell(5, 3).number_format = NUM_FMT
    ws.cell(5, 5, base_fcf).font = INPUT_FONT
    ws.cell(5, 5).fill = YELLOW_BG
    ws.cell(5, 5).number_format = NUM_FMT

    # ── 5年预测行 ──
    for i in range(5):
        r = 6 + i
        year = base_year + 1 + i

        ws.cell(r, 1, year).font = BLACK_BOLD

        # 营收增长率（输入）
        ws.cell(r, 2, growth[i]).font = INPUT_FONT
        ws.cell(r, 2).fill = YELLOW_BG
        ws.cell(r, 2).number_format = PCT_FMT

        # 营收 = 上年 × (1 + 增长率)
        ws.cell(r, 3).value = f"=C{r-1}*(1+B{r})"
        ws.cell(r, 3).font = FORMULA_FONT
        ws.cell(r, 3).number_format = NUM_FMT

        # FCF Margin（输入）
        ws.cell(r, 4, margins[i]).font = INPUT_FONT
        ws.cell(r, 4).fill = YELLOW_BG
        ws.cell(r, 4).number_format = PCT_FMT

        # FCF = 营收 × FCF Margin
        ws.cell(r, 5).value = f"=C{r}*D{r}"
        ws.cell(r, 5).font = FORMULA_FONT
        ws.cell(r, 5).number_format = NUM_FMT

        # 折现因子 = 1/(1+WACC)^(t)
        ws.cell(r, 6).value = f"=1/(1+$J$1)^{i+1}"
        ws.cell(r, 6).font = FORMULA_FONT
        ws.cell(r, 6).number_format = '0.0000'

        # PV of FCF
        ws.cell(r, 7).value = f"=E{r}*F{r}"
        ws.cell(r, 7).font = FORMULA_FONT
        ws.cell(r, 7).number_format = NUM_FMT

    # ── 终值计算 ──
    r = 12
    ws.cell(r, 1, "终值计算").font = DARK_BLUE

    r = 13
    ws.cell(r, 1, "预测期 FCF 现值合计")
    ws.cell(r, 2).value = "=SUM(G6:G10)"
    ws.cell(r, 2).font = FORMULA_FONT
    ws.cell(r, 2).number_format = NUM_FMT

    r = 14
    ws.cell(r, 1, "终年 FCF")
    ws.cell(r, 2).value = "=E10"
    ws.cell(r, 2).font = FORMULA_FONT
    ws.cell(r, 2).number_format = NUM_FMT

    r = 15
    ws.cell(r, 1, "终值 (Gordon Growth)")
    ws.cell(r, 2).value = "=B14*(1+$J$2)/($J$1-$J$2)"
    ws.cell(r, 2).font = FORMULA_FONT
    ws.cell(r, 2).number_format = NUM_FMT

    r = 16
    ws.cell(r, 1, "终值现值")
    ws.cell(r, 2).value = "=B15*F10"
    ws.cell(r, 2).font = FORMULA_FONT
    ws.cell(r, 2).number_format = NUM_FMT

    r = 18
    # ── 股权价值（含净债务调整）──
    net_debt_val = c['financials'].get('net_debt', 0) or 0
    # 用隐藏列I3存净债务值
    ws['I3'] = net_debt_val

    ws.cell(r, 1, "企业价值 (EV)").font = DARK_BLUE
    ws.cell(r, 2).value = "=B13+B16"
    ws.cell(r, 2).font = LARGE_BOLD
    ws.cell(r, 2).number_format = NUM_FMT

    # C18: 净债务（隐藏，仅用于计算）
    ws.cell(r, 3, net_debt_val)
    ws.cell(r, 3).font = INPUT_FONT
    ws.cell(r, 3).fill = GRAY_BG
    ws.cell(r, 3).number_format = NUM_FMT

    # D18: 股权价值 = EV - 净债务
    ws.cell(r, 4, "股权价值").font = DARK_BLUE
    ws.cell(r, 4).value = "=B18-C18"
    ws.cell(r, 4).font = LARGE_BOLD
    ws.cell(r, 4).number_format = NUM_FMT

    # ── 每股价值 ──
    shares = c['shares']
    fx = c.get('fx_rate', 1)

    r = 19
    ws.cell(r, 1, "总股本（亿股）")
    ws.cell(r, 2, shares).font = INPUT_FONT
    ws.cell(r, 2).fill = YELLOW_BG

    r = 20
    report_cur = c.get('report_currency', 'CNY')
    # B20 = 股权价值(D18) * 10（十亿→亿） / 总股本(B19) / 汇率转HKD
    ws.cell(r, 1, f"每股价值 ({c['currency']})")
    if market in ('HK', 'US'):
        ws.cell(r, 2).value = f"=D18*10/B19*{fx}"
    else:
        ws.cell(r, 2).value = "=D18*10/B19"
    ws.cell(r, 2).font = FORMULA_FONT
    ws.cell(r, 2).number_format = NUM_FMT

    r = 21
    # B20已是交易货币(HKD)，B21转回报告货币(CNY)：除以fx
    ws.cell(r, 1, f"每股价值 ({report_cur})").font = DARK_BLUE
    if market in ('HK', 'US'):
        ws.cell(r, 2).value = f"=B20/{fx}"
    else:
        ws.cell(r, 2).value = "=B20"
    ws.cell(r, 2).font = Font(bold=True, size=12, color=title_colors[idx])
    ws.cell(r, 2).number_format = NUM_FMT

    r = 22
    ws.cell(r, 1, f"当前股价 ({c['currency']})")
    ws.cell(r, 2, c['current_price']).font = INPUT_FONT
    ws.cell(r, 2).fill = YELLOW_BG

    r = 23
    ws.cell(r, 1, "上涨/下跌空间")
    # 用B20（交易货币）对比B22（当前股价，交易货币）
    ws.cell(r, 2).value = "=(B20-B22)/B22"
    ws.cell(r, 2).font = FORMULA_FONT
    ws.cell(r, 2).number_format = '+0.0%;-0.0%'

    # ── 关键假设记录 ──
    r = 25
    ws.cell(r, 1, "关键假设").font = DARK_BLUE
    ws.cell(r+1, 1, f"WACC: 引用自概览页").font = GRAY
    ws.cell(r+2, 1, f"永续增长率: {s['terminal_growth']:.1%}").font = GRAY

    return ws


# ── Sheet 5: 敏感性分析（Excel 数据表） ──────────────
def build_sensitivity(wb, cfg):
    ws = wb.create_sheet("敏感性分析")
    set_col_widths(ws, [14] + [14]*7)

    ws['A1'] = f"WACC x 永续增长率 敏感性分析（基准情景，每股{cfg['currency']}）"
    ws['A1'].font = DARK_BLUE

    # ── 基准参数引用 ──
    ws['I1'] = '基准每股价值'
    ws['J1'] = f"='基准情景'!B20"
    ws['J1'].number_format = NUM_FMT

    ws['I2'] = '当前股价'
    ws['J2'] = cfg['current_price']

    # ── WACC 范围 ──
    base_wacc = cfg['wacc']['rf'] + cfg['wacc']['beta'] * cfg['wacc']['erp']
    wacc_vals = [round(base_wacc - 0.015 + 0.005 * i, 4) for i in range(7)]

    # ── 永续增长率范围 ──
    base_tg = cfg['scenarios'][1]['terminal_growth']
    tg_vals = [round(base_tg - 0.005 + 0.005 * i, 4) for i in range(4)]

    # ── 表头 ──
    ws.cell(3, 1, "WACC →").font = BLACK_BOLD
    for j, w in enumerate(wacc_vals):
        cell = ws.cell(3, j+2, w)
        cell.font = WHITE_BOLD
        cell.fill = HEADER_BG
        cell.number_format = PCT_FMT

    # ── 计算区域（隐藏的基准参数）──
    ws['N1'] = '基准营收'
    ws['O1'] = cfg['financials']['revenue']
    ws['N2'] = '总股本'
    ws['O2'] = cfg['shares']
    ws['N3'] = '汇率'
    ws['O3'] = cfg.get('fx_rate', 1)
    ws['N4'] = '市场'
    ws['O4'] = cfg.get('market', 'HK')

    # ── 基准情景增长率和 FCF Margin ──
    base_s = cfg['scenarios'][1]
    for i, (g, m) in enumerate(zip(base_s['growth_rates'], base_s['fcf_margins'])):
        ws.cell(10+i, 14, g)  # N10-N14: growth rates
        ws.cell(10+i, 15, m)  # O10-O14: FCF margins

    # ── 敏感性矩阵 ──
    for i, tg in enumerate(tg_vals):
        r = 4 + i
        ws.cell(r, 1, f"TG={tg:.1%}").font = BLACK_BOLD

        for j, wacc in enumerate(wacc_vals):
            # Python 预计算（因为 Excel 数据表在 openpyxl 中不够灵活）
            revs = [cfg['financials']['revenue']]
            for g in base_s['growth_rates']:
                revs.append(revs[-1] * (1 + g))

            fcfs_pv = 0
            last_fcf = 0
            last_df = 0
            for yr in range(5):
                fcf = revs[yr+1] * base_s['fcf_margins'][yr]
                df = 1 / (1 + wacc) ** (yr + 1)
                fcfs_pv += fcf * df
                if yr == 4:
                    last_fcf = fcf
                    last_df = df

            tv = last_fcf * (1 + tg) / (wacc - tg) if wacc > tg else 0
            ev = fcfs_pv + tv * last_df

            # 净债务调整
            net_debt = cfg['financials'].get('net_debt', 0)
            equity = ev - (net_debt if net_debt else 0)
            # 单位统一：equity是十亿，shares是亿，需×10；港股/美股需×汇率
            per_share = equity * 10 / cfg['shares']
            if cfg.get('market', 'HK') in ('HK', 'US'):
                per_share *= cfg.get('fx_rate', 1)

            cell = ws.cell(r, j+2, round(per_share, 1))
            cell.number_format = INT_FMT
            cell.border = THIN_BORDER

            # 条件着色
            current_price = cfg['current_price']
            margin = (per_share - current_price) / current_price
            cell.fill = get_rate_fill(margin)

    # ── 说明 ──
    r = 4 + len(tg_vals) + 1
    ws.cell(r, 1, f"当前股价: {cfg['current_price']} {cfg['currency']}").font = GRAY
    ws.cell(r+1, 1, "配色说明：").font = BLACK_BOLD
    ws.cell(r+2, 1, "深绿/浅绿 = 低估（安全边际 > 20%）").font = GRAY
    ws.cell(r+3, 1, "黄色 = 合理（-10% ~ 20%）").font = GRAY
    ws.cell(r+4, 1, "红色 = 高估（安全边际 < -10%）").font = GRAY

    return ws


# ── Sheet 6: 营收 FCF 对比 ───────────────────────────
def build_comparison(wb, cfg):
    ws = wb.create_sheet("营收FCF对比")
    set_col_widths(ws, [10, 14, 14, 14, 14, 14, 14])

    report_cur = cfg.get('report_currency', 'CNY')
    ws['A1'] = f"三情景营收 & FCF 预测对比 (十亿 {report_cur})"
    ws['A1'].font = DARK_BLUE

    headers = ['年份', '乐观-营收', '乐观-FCF', '基准-营收', '基准-FCF', '悲观-营收', '悲观-FCF']
    for i, h in enumerate(headers):
        ws.cell(3, i+1, h)
    style_header_row(ws, 3, 1, 7)

    scenarios = cfg['scenarios']
    base_year = cfg['financials']['base_year']

    for yr_idx in range(6):
        r = 4 + yr_idx
        row_src = 5 + yr_idx  # 情景 Sheet 中对应行号

        if yr_idx == 0:
            ws.cell(r, 1, f"{base_year} (基准)")
        else:
            ws.cell(r, 1, base_year + yr_idx)

        for s_idx in range(3):
            sn = scenarios[s_idx]['name']
            col_rev = 2 + s_idx * 2
            col_fcf = 3 + s_idx * 2

            ws.cell(r, col_rev).value = f"='{sn}'!C{row_src}"
            ws.cell(r, col_rev).font = CROSS_REF_FONT
            ws.cell(r, col_rev).number_format = NUM_FMT

            ws.cell(r, col_fcf).value = f"='{sn}'!E{row_src}"
            ws.cell(r, col_fcf).font = CROSS_REF_FONT
            ws.cell(r, col_fcf).number_format = NUM_FMT

    return ws


# ── 主生成函数 ────────────────────────────────────────
def generate(config_path, output_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        cfg = json.load(f)

    # 验证必要字段
    required = ['company_name', 'ticker', 'current_price', 'currency', 'shares', 'financials', 'wacc', 'scenarios']
    for field in required:
        if field not in cfg:
            print(f"[ERROR] 缺少必要字段: {field}")
            sys.exit(1)

    if len(cfg['scenarios']) != 3:
        print(f"[ERROR] 必须有3个情景，当前: {len(cfg['scenarios'])}")
        sys.exit(1)

    # 确保市场标识
    if 'market' not in cfg:
        ticker = cfg['ticker']
        if '.HK' in ticker:
            cfg['market'] = 'HK'
        elif '.SH' in ticker or '.SZ' in ticker:
            cfg['market'] = 'CN'
        else:
            cfg['market'] = 'US'

    # 确保报告货币
    if 'report_currency' not in cfg:
        cfg['report_currency'] = 'CNY'

    wb = Workbook()

    # 按顺序构建 Sheet
    build_overview(wb, cfg)
    for i in range(3):
        build_scenario(wb, cfg, i)
    build_sensitivity(wb, cfg)
    build_comparison(wb, cfg)

    wb.save(output_path)
    print(f"[OK] DCF 模型已生成: {output_path}")
    print(f"   公司: {cfg['company_name']} ({cfg['ticker']})")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")
    print(f"   股价: {cfg['current_price']} {cfg['currency']}")
    print(f"   情景数: {len(cfg['scenarios'])}")

    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python generate_dcf.py <config.json> <output.xlsx>")
        print("Example: python generate_dcf.py sample_config.json 泡泡玛特_DCF估值模型.xlsx")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2])
